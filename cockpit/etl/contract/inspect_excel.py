"""Inspeciona estrutura do Excel (abas e cabeçalhos)."""
import sys

path = r"c:\Users\Vitor A. Tito\Documents\GPTO\GSN\2026\wms\cockpit\etl\contract\VOLUME COMERCIAL 10.12.xlsx"

# Tentar xlsx com openpyxl (read_only=False para evitar problemas)
try:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    print("=== ABAS (openpyxl) ===")
    for i, name in enumerate(wb.sheetnames):
        print(f"  {i+1}. {name}")
    for name in wb.sheetnames[:15]:  # primeiras 15 abas
        ws = wb[name]
        rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
        headers = rows[0] if rows else []
        print(f"\n=== ABA: {name} ===")
        print("  Cabeçalhos:", [str(h)[:50] if h is not None else "" for h in (headers or [])])
        if len(rows) > 1 and rows[1]:
            print("  Amostra linha 2:", [str(v)[:30] if v is not None else "" for v in rows[1][:12]])
    wb.close()
    sys.exit(0)
except Exception as e:
    print("openpyxl falhou:", e, file=sys.stderr)

# Tentar com pandas (pode usar engine diferente)
try:
    import pandas as pd
    xl = pd.ExcelFile(path)
    print("\n=== ABAS (pandas) ===")
    for i, name in enumerate(xl.sheet_names):
        print(f"  {i+1}. {name}")
    for name in xl.sheet_names[:15]:
        df = pd.read_excel(path, sheet_name=name, nrows=2, header=0)
        print(f"\n=== ABA: {name} ===")
        print("  Colunas:", list(df.columns))
        if len(df) > 0:
            print("  Amostra:", df.iloc[0].astype(str).str[:30].tolist()[:12])
except Exception as e:
    print("pandas falhou:", e, file=sys.stderr)
    sys.exit(1)
