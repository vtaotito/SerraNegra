"""
Extrai estrutura do Excel (abas, colunas, tipos inferidos, amostras) para JSON.
Uso: python extract_excel_structure.py
"""
import json
import openpyxl
from pathlib import Path
from datetime import datetime, date

EXCEL_PATH = Path(__file__).parent / "VOLUME COMERCIAL 10.12.xlsx"
OUT_JSON = Path(__file__).parent / "excel_structure.json"
OUT_UI = Path(__file__).parent / "excel_structure_for_ui.json"
MAX_SAMPLE_ROWS = 3
MAX_COLS_DISPLAY = 30


def infer_type(v):
    if v is None: return "null"
    if isinstance(v, bool): return "boolean"
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v == int(v): return "integer"
        return "number"
    if isinstance(v, (datetime, date)): return "date"
    if isinstance(v, str): return "string"
    return "string"


def extract():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    result = {
        "source": EXCEL_PATH.name,
        "extractedAt": datetime.now().isoformat(),
        "sheets": []
    }
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result["sheets"].append({
                "name": sheet_name,
                "rows": 0,
                "columns": 0,
                "headers": [],
                "sampleRows": [],
                "columnTypes": []
            })
            continue
        headers = [str(h).strip() if h is not None else f"Col_{i}" for i, h in enumerate(rows[0])]
        ncols = len(headers)
        data_rows = rows[1 : 1 + MAX_SAMPLE_ROWS]
        sample_rows = []
        for row in data_rows:
            arr = list(row)[:MAX_COLS_DISPLAY]
            sample_rows.append([v if not isinstance(v, (datetime, date)) else v.isoformat() for v in arr])
        # infer types from first data rows
        col_types = []
        for c in range(min(ncols, MAX_COLS_DISPLAY)):
            types_seen = set()
            for row in data_rows:
                if c < len(row) and row[c] is not None:
                    types_seen.add(infer_type(row[c]))
            if not types_seen:
                col_types.append("string")
            elif "number" in types_seen:
                col_types.append("number")
            elif "integer" in types_seen and "number" not in types_seen:
                col_types.append("integer")
            elif "date" in types_seen:
                col_types.append("date")
            else:
                col_types.append("string")
        result["sheets"].append({
            "name": sheet_name,
            "rows": len(rows) - 1,
            "columns": ncols,
            "headers": headers[:MAX_COLS_DISPLAY],
            "columnTypes": col_types,
            "sampleRows": sample_rows
        })
    wb.close()
    return result


if __name__ == "__main__":
    data = extract()
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"Written {OUT_JSON}")
    # Versão simplificada para UI (catálogo de abas + colunas + tipos)
    ui_data = {
        "sheets": [
            {
                "id": i + 1,
                "name": s["name"],
                "rowCount": s["rows"],
                "columns": [
                    {"key": h or f"col_{j}", "label": h or f"Coluna {j+1}", "type": s["columnTypes"][j] if j < len(s["columnTypes"]) else "string"}
                    for j, h in enumerate(s["headers"])
                ],
                "sampleRows": s["sampleRows"]
            }
            for i, s in enumerate(data["sheets"])
        ]
    }
    with open(OUT_UI, "w", encoding="utf-8") as f:
        json.dump(ui_data, f, ensure_ascii=False, indent=2)
    print(f"Written {OUT_UI}")
    print("Sheets:", [s["name"] for s in data["sheets"]])
