"""
Deep analysis of VOLUME COMERCIAL 10.12.xlsx (password-protected).
Extracts: sheet names, real headers (handling pivot tables),
sample data, KPI-relevant aggregations, date ranges, and unique values for filters.
"""
import json, sys, os
from datetime import datetime

try:
    import msoffcrypto
    import io
    HAVE_CRYPTO = True
except ImportError:
    HAVE_CRYPTO = False

import openpyxl

EXCEL_PATH = r"c:\Users\Vitor A. Tito\Documents\GPTO\GSN\2026\VOLUME COMERCIAL 10.12.xlsx"
PASSWORD = "GSN9653"
OUTPUT_PATH = r"c:\Users\Vitor A. Tito\Documents\GPTO\GSN\2026\wms\cockpit\etl\ingest\excel_deep_analysis.json"

def open_workbook(path, password):
    return openpyxl.load_workbook(path, data_only=True, read_only=True)

def cell_val(cell):
    v = cell.value
    if isinstance(v, datetime):
        return v.isoformat()
    return v

def analyze_sheet(ws, name, max_rows=500):
    rows_data = []
    for i, row in enumerate(ws.iter_rows(max_row=max_rows)):
        if i > max_rows:
            break
        rows_data.append([cell_val(c) for c in row])

    if not rows_data:
        return {"name": name, "rows": 0, "columns": 0, "headers": [], "sampleRows": []}

    headers = rows_data[0] if rows_data else []
    col_count = len(headers)

    all_none_headers = all(h is None for h in headers)
    real_headers = []
    for i, h in enumerate(headers):
        if h is not None:
            real_headers.append(str(h))
        else:
            real_headers.append(f"Col_{i}")

    num_cols = {}
    date_cols = {}
    str_cols = {}
    for row in rows_data[1:min(50, len(rows_data))]:
        for j, v in enumerate(row):
            if v is None:
                continue
            if isinstance(v, (int, float)):
                num_cols[j] = num_cols.get(j, 0) + 1
            elif isinstance(v, str) and "T" in v and "-" in v:
                date_cols[j] = date_cols.get(j, 0) + 1
            elif isinstance(v, str):
                str_cols[j] = str_cols.get(j, 0) + 1

    col_types = []
    for j in range(col_count):
        if j in date_cols and date_cols[j] > num_cols.get(j, 0):
            col_types.append("date")
        elif j in num_cols and num_cols[j] > str_cols.get(j, 0):
            col_types.append("number")
        else:
            col_types.append("string")

    unique_values = {}
    for j in range(min(col_count, 30)):
        vals = set()
        for row in rows_data[1:min(200, len(rows_data))]:
            if j < len(row) and row[j] is not None:
                vals.add(str(row[j]))
        if 1 < len(vals) <= 30:
            unique_values[real_headers[j]] = sorted(list(vals))[:20]

    numeric_summaries = {}
    for j in range(min(col_count, 30)):
        if col_types[j] == "number":
            vals = []
            for row in rows_data[1:]:
                if j < len(row) and isinstance(row[j], (int, float)):
                    vals.append(row[j])
            if vals:
                numeric_summaries[real_headers[j]] = {
                    "min": round(min(vals), 2),
                    "max": round(max(vals), 2),
                    "sum": round(sum(vals), 2),
                    "avg": round(sum(vals)/len(vals), 2),
                    "count": len(vals),
                }

    return {
        "name": name,
        "totalRows": ws.max_row if hasattr(ws, 'max_row') else len(rows_data),
        "totalColumns": ws.max_column if hasattr(ws, 'max_column') else col_count,
        "analyzedRows": len(rows_data),
        "headers": real_headers,
        "columnTypes": col_types,
        "sampleRows": rows_data[1:6],
        "uniqueFilterValues": unique_values,
        "numericSummaries": numeric_summaries,
    }

def main():
    print(f"Opening {EXCEL_PATH}...")
    wb = open_workbook(EXCEL_PATH, PASSWORD)
    print(f"Sheets: {wb.sheetnames}")

    results = {
        "source": os.path.basename(EXCEL_PATH),
        "extractedAt": datetime.now().isoformat(),
        "password": "(protected)",
        "sheets": []
    }

    for sname in wb.sheetnames:
        print(f"  Analyzing: {sname}...")
        ws = wb[sname]
        info = analyze_sheet(ws, sname)
        results["sheets"].append(info)
        print(f"    > {info.get('totalRows', '?')} rows, {info.get('totalColumns', '?')} cols, {len(info.get('numericSummaries', {}))} numeric cols")

    wb.close()

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2, default=str)
    print(f"\nSaved to {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
